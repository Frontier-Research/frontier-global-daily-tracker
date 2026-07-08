"""
fetch.py — the twice-a-day pipeline.

Flow:
  1. Load the accumulated history from data/store.xlsx (one tab per instrument).
  2. For each instrument, fetch fresh data from its source.
       - New dates are appended.
       - The most recent stored date may be refreshed (provisional -> final,
         or an intraday move for the 24/7 markets).
       - If the source has nothing newer (weekend / holiday), nothing is appended
         and the instrument is flagged with a note.
       - If a fetch fails, the stored history is kept untouched.
  3. Recompute today / prior / 2-day / 1-week closes and the change columns from
     the *merged* history (so charts and the board use the long memory, not just
     the latest pull).
  4. Write everything back atomically:
       data/store.xlsx           — the master multi-tab store (also a download)
       data/series/<id>.json     — per-instrument daily series for the charts
       data/latest.json          — the board snapshot the dashboard reads first
       data/meta.json            — build timestamps (UTC + Sri Lanka time)

The store is the single accumulating record: even where a source only offers a
limited window, the store grows one day at a time on every run.
"""

from __future__ import annotations

import json
import os
import re
from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook

from sources import INSTRUMENTS, Instrument

# Paths are resolved relative to the repo root (parent of scripts/).
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(ROOT, "data")
SERIES = os.path.join(DATA, "series")
STORE_XLSX = os.path.join(DATA, "store.xlsx")

SLT = ZoneInfo("Asia/Colombo")
SERIES_CAP = 2000  # max daily points kept in each chart JSON (~8 years)

_SHEET_SAFE = re.compile(r"[\[\]\*\?/\\:]")


# --------------------------------------------------------------------------- #
# IO helpers                                                                   #
# --------------------------------------------------------------------------- #
def _atomic_write_json(path: str, payload) -> None:
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, separators=(",", ":"))
    os.replace(tmp, path)


def _sheet_name(name: str) -> str:
    return _SHEET_SAFE.sub("", name)[:31]


# --------------------------------------------------------------------------- #
# Store load / save                                                            #
# --------------------------------------------------------------------------- #
def load_store() -> dict[str, dict[str, float]]:
    """Return {instrument_id: {date: close}} from the existing workbook, if any."""
    store: dict[str, dict[str, float]] = {inst.id: {} for inst in INSTRUMENTS}
    if not os.path.exists(STORE_XLSX):
        return store
    wb = load_workbook(STORE_XLSX, read_only=True, data_only=True)
    by_sheet = {_sheet_name(inst.name): inst.id for inst in INSTRUMENTS}
    for ws in wb.worksheets:
        inst_id = by_sheet.get(ws.title)
        if not inst_id:
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None or row[1] is None:
                continue
            date = str(row[0])[:10]
            try:
                store[inst_id][date] = float(row[1])
            except (TypeError, ValueError):
                continue
    wb.close()
    return store


def save_store(store: dict[str, dict[str, float]]) -> None:
    """Rewrite the whole workbook, one tab per instrument, sorted by date."""
    wb = Workbook()
    wb.remove(wb.active)
    for inst in INSTRUMENTS:
        ws = wb.create_sheet(title=_sheet_name(inst.name))
        ws.append(["Date", "Close", "Chg", "Chg %", "Source"])
        series = sorted(store[inst.id].items())
        prev = None
        for date, close in series:
            if prev is None:
                chg = chg_pct = None
            else:
                chg = round(close - prev, 6)
                chg_pct = round((close - prev) / prev * 100, 4) if prev else None
            ws.append([date, round(close, 6), chg, chg_pct, inst.source])
            prev = close
        ws.freeze_panes = "A2"
    tmp = STORE_XLSX + ".tmp"
    wb.save(tmp)
    os.replace(tmp, STORE_XLSX)


# --------------------------------------------------------------------------- #
# Merge + derive                                                               #
# --------------------------------------------------------------------------- #
def merge(existing: dict[str, float], new_points: list[tuple[str, float]]):
    """
    Fold freshly fetched points into the stored series.
    Returns (merged_dict, updated_bool, note).
    """
    merged = dict(existing)
    last_existing = max(existing) if existing else None
    added_new_date = False
    refreshed_latest = False

    for date, value in new_points:
        if date not in merged:
            merged[date] = value
            if last_existing is None or date > last_existing:
                added_new_date = True
        elif last_existing is not None and date == last_existing and merged[date] != value:
            merged[date] = value           # refine provisional / intraday value
            refreshed_latest = True

    updated = added_new_date or refreshed_latest
    new_max = max(d for d, _ in new_points)
    if not added_new_date and last_existing is not None and new_max <= last_existing:
        note = f"No new close since {last_existing}"
    else:
        note = ""
    return merged, updated, note


def _nearest_on_or_before(series: list[tuple[str, float]], target: str):
    """Last (date, value) with date <= target, or None."""
    hit = None
    for date, value in series:
        if date <= target:
            hit = (date, value)
        else:
            break
    return hit


def derive(series_map: dict[str, float]) -> dict:
    """Compute the board snapshot for one instrument from its full history."""
    series = sorted(series_map.items())
    n = len(series)
    out = {
        "today": None, "prior": None, "twoday": None, "week": None,
        "chg": None, "chg_pct": None, "twoday_chg_pct": None, "week_chg_pct": None,
        "last_date": None, "spark": [],
    }
    if n == 0:
        return out

    today_date, today = series[-1]
    out["today"] = today
    out["last_date"] = today_date
    out["spark"] = [round(v, 6) for _, v in series[-30:]]

    if n >= 2:
        out["prior"] = series[-2][1]
    if n >= 3:
        out["twoday"] = series[-3][1]

    wk_target = (datetime.strptime(today_date, "%Y-%m-%d") - timedelta(days=7)).strftime("%Y-%m-%d")
    wk = _nearest_on_or_before(series[:-1], wk_target)
    if wk:
        out["week"] = wk[1]

    def pct(a, b):
        return round((a - b) / b * 100, 4) if b else None

    if out["prior"] is not None:
        out["chg"] = round(today - out["prior"], 6)
        out["chg_pct"] = pct(today, out["prior"])
    if out["twoday"] is not None:
        out["twoday_chg_pct"] = pct(today, out["twoday"])
    if out["week"] is not None:
        out["week_chg_pct"] = pct(today, out["week"])
    return out


# --------------------------------------------------------------------------- #
# Main                                                                         #
# --------------------------------------------------------------------------- #
def main() -> None:
    os.makedirs(SERIES, exist_ok=True)
    store = load_store()

    now_utc = datetime.now(timezone.utc)
    now_slt = now_utc.astimezone(SLT)
    rows = []
    notes: list[str] = []

    for inst in INSTRUMENTS:
        note = ""
        try:
            points = inst.fetch()
            store[inst.id], _updated, note = merge(store[inst.id], points)
        except Exception as exc:  # noqa: BLE001 - keep other instruments running
            note = f"Source unavailable this run ({type(exc).__name__})"
            notes.append(f"{inst.name}: {note}")

        metrics = derive(store[inst.id])

        # Per-instrument chart series (capped, ascending).
        series_sorted = sorted(store[inst.id].items())[-SERIES_CAP:]
        _atomic_write_json(os.path.join(SERIES, f"{inst.id}.json"), {
            "id": inst.id, "name": inst.name, "group": inst.group, "color": inst.color,
            "unit": inst.unit, "decimals": inst.decimals, "source": inst.source,
            "points": [[d, round(v, 6)] for d, v in series_sorted],
        })

        rows.append({
            "id": inst.id, "name": inst.name, "group": inst.group, "color": inst.color,
            "unit": inst.unit, "decimals": inst.decimals, "source": inst.source,
            "order": inst.order,
            **metrics,
            "note": note,
        })

    save_store(store)

    generated = {
        "generated_utc": now_utc.strftime("%Y-%m-%dT%H:%M:%SZ"),
        "generated_slt": now_slt.strftime("%Y-%m-%d %H:%M"),
        "generated_slt_long": now_slt.strftime("%A, %d %B %Y · %H:%M"),
    }
    _atomic_write_json(os.path.join(DATA, "latest.json"), {**generated, "rows": rows})
    _atomic_write_json(os.path.join(DATA, "meta.json"), {
        **generated,
        "instruments": [
            {"id": r["id"], "name": r["name"], "last_date": r["last_date"],
             "note": r["note"], "source": r["source"]}
            for r in rows
        ],
        "notes": notes,
    })
    print(f"Done. {len(rows)} instruments. Notes: {len(notes)}")


if __name__ == "__main__":
    main()
